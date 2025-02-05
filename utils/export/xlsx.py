from openpyxl import Workbook
import pandas as pd
from tqdm import tqdm
from tqdm.contrib.logging import logging_redirect_tqdm
from collections import defaultdict
import logging
import os
import datetime
from utils import DESC_WIDTH

class xlsxWriter:
    """
    Exports traffic counting results to Excel format.
    
    Creates detailed reports including:
    - Crossing timestamps
    - Vehicle classifications
    - Direction information
    - Confidence scores
    """
    def __init__(self, progress_callback=None):
        """
        Args:
            progress_callback: Optional callback function to report export progress
        """
        self.progress_callback = progress_callback
        # Create a new workbook
        self.workbook = Workbook()
        # Create a new sheet
        self.sheet = self.workbook.active

    def __prepare_data(self, data_manager):
        self.write_data = []

        for obj_id, crossings in data_manager.CROSSED.items():
            track_analysis = data_manager.TRACK_ANALYSIS.get(obj_id, {})
            
            for (frame, cls, direction, _, final_conf, stats) in crossings:
                timestamp = frame / data_manager.fps
                crossing_time = data_manager.start_datetime + datetime.timedelta(seconds=timestamp)
                
                # Basic data
                row_data = [
                    data_manager.site_location,
                    crossing_time.date(),
                    crossing_time.date() - datetime.timedelta(days=crossing_time.date().weekday()),
                    crossing_time.strftime("%A"),
                    crossing_time.time(),
                    f"{crossing_time.hour}:{(crossing_time.minute // 15) * 15:02d}",
                    f"{crossing_time.hour}:00",
                    direction,
                    data_manager.names[cls],
                    obj_id,
                    f"{final_conf:.3f}"  # Overall confidence
                ]

                # Add class statistics
                if stats:
                    cls_stats = stats[cls]
                    row_data.extend([
                        cls_stats['count'],  # Number of frames as this class
                        f"{cls_stats['total_conf']/cls_stats['count']:.3f}",  # Avg confidence
                        cls_stats['max_consecutive']  # Longest consecutive detection
                    ])
                else:
                    row_data.extend([0, "0.000", 0])

                self.write_data.append(row_data)

    def write_to_excel(self, export_path_excel, data_manager, progress_var=None):
        self.progress = progress_var
        self.__prepare_data(data_manager)
        # Write the headers into the columns
        headers = [
            "Site", "Date", "First day of Week", "Week day",
            "Time of crossing", "15 Min Interval", "Hour Interval",
            "Direction", "Class", "ID", "Track Confidence",
            "Frame Count", "Average Confidence", "Max Consecutive Frames"
        ]
        
        self.sheet.append(headers)

        # Write the data
        length, prog_count = len(self.write_data), 0
        self.console_progress = tqdm(total=length, desc=f'{"Writing xlsx report":<{DESC_WIDTH}}', unit="rows", dynamic_ncols=True)
        with logging_redirect_tqdm():
            for row in self.write_data:
                self.sheet.append(row)
                self.console_progress.update(1)
                prog_count += 1 
                if self.progress is not None : self.progress = prog_count// length
                if self.progress_callback:
                        progress_percentage = int((prog_count / length) * 100)
                        self.progress_callback(progress_percentage)

        # Check if same file exists and enumerate names if it does
        base, extension = os.path.splitext(export_path_excel)
        counter = 1
        new_export_path_excel = export_path_excel

        while os.path.exists(new_export_path_excel):
            new_export_path_excel = f"{base}_{counter}{extension}"
            counter += 1

        export_path_excel = new_export_path_excel

        # Save the workbook
        self.workbook.save(export_path_excel)

        self.console_progress.close()
        return export_path_excel
    
class xlsxCompiler:
    """
    Combines multiple Excel reports into a single consolidated report.
    
    Aggregates counting data from multiple files while maintaining:
    - Site information
    - Time intervals
    - Vehicle classifications
    - Directional data
    """
    def __init__(self, folder_path=None, file_paths=None):
        self.folder_path = folder_path
        self.file_paths = file_paths if file_paths else []
        self.object_data = {} 
        self.compiled_data = {} 

    def read_files(self):
        if self.folder_path:
            files = [os.path.join(self.folder_path, f) for f in os.listdir(self.folder_path) if f.endswith('.xlsx')]
        else:
            files = self.file_paths

        for file in files:
            self.extract_data(file)

    def extract_data(self, file_path):
        xls = pd.ExcelFile(file_path)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            for _, row in df.iterrows():
                obj_id = row['ID']
                site_location = row['Site']
                date = row['Date']
                vehicle_type = row['Class']
                direction = row['Direction']
                time = row['Time of crossing']
                interval_15_min = row['15 Min Interval']
                interval_hr = row['Hour Interval']

                if obj_id not in self.object_data:
                    self.object_data[obj_id] = {
                        'Site': site_location,
                        'Date': date,
                        'Vehicle Type': vehicle_type,
                        'Directions': set(),
                        'Time of last Crossing': time,
                        '15 Min Interval': interval_15_min,
                        'Hour Interval': interval_hr
                    }
                self.object_data[obj_id]['Directions'].add(direction)
                self.object_data[obj_id]['Time of Crossing'] = time
                self.object_data[obj_id]['15 Min Interval'] = interval_15_min
                self.object_data[obj_id]['Hour Interval'] = interval_hr

    def precompile(self):
        # Process object_data to concatenate directions
        for obj_id, data in self.object_data.items():
            site_location = data['Site']
            date = data['Date'].strftime('%Y-%m-%d')
            vehicle_type = data['Vehicle Type']
            directions = " - ".join(sorted(data['Directions']))
            interval_15_min = data['15 Min Interval']
            interval_hr = data['Hour Interval']
            time = data['Time of last Crossing']

            key = (site_location, date, vehicle_type, directions, interval_15_min, interval_hr)
            if key not in self.compiled_data:
                self.compiled_data[key] = 0
            self.compiled_data[key] += 1

    def write_compiled_data(self, output_path):
        workbook = Workbook()
        sheet = workbook.active

        headers = ['Site/Location', 'Date', 'Vehicle Type', 'Direction', '15 Min Interval', 'Hour Interval', 'Total Count']
        sheet.append(headers)

        for key, count in self.compiled_data.items():
            sheet.append(list(key) + [count])

        workbook.save(output_path)

    def compile(self, output_path):
        self.read_files()
        self.precompile()
        self.write_compiled_data(output_path)

class StreetCountCompiler:
    def __init__(self, file_paths, site_location, timezone):
        self.file_paths = file_paths
        self.site_location = site_location
        self.timezone = timezone
        self.object_data = defaultdict(list)  # To store raw data
        self.compiled_data = defaultdict(int)  # To store aggregated counts
        
    def extract_data(self):
        """
        Extracts data from each CSV file and stores it in object_data.
        """
        for file_path in self.file_paths:
            try:
                df = pd.read_csv(file_path, header=None, names=["Timestamp", "Direction", "Vehicle Type"])
                
                # Parse the Timestamp to datetime objects and adjust for timezone
                df['Timestamp'] = pd.to_datetime(df['Timestamp'], format='%Y-%m-%dT%H:%M:%S.%fZ').dt.tz_localize('UTC').dt.tz_convert(self.timezone)
                
                # Append to object_data
                for _, row in df.iterrows():
                    self.object_data["records"].append({
                        "timestamp": row["Timestamp"],
                        "direction": row["Direction"],
                        "vehicle_type": row["Vehicle Type"]
                    })
            except Exception as e:
                logging.error(f"Error processing file {file_path}: {e}")
    
    def precompile(self):
        """
        Aggregates the extracted data into compiled_data.
        """

        for record in self.object_data["records"]:
            date = record["timestamp"].date()
            hour = record["timestamp"].hour
            minute = record["timestamp"].minute
            # Calculate 15-minute interval
            interval_15 = (minute // 15) * 15
            # Format interval strings
            interval_15_str = f"{interval_15:02d}:00 - {interval_15 + 15:02d}:00"
            interval_hour_str = f"{hour:02d}:00 - {hour:02d}:59"

            key = (
                self.site_location,
                date.strftime('%Y-%m-%d'),
                record["vehicle_type"],
                record["direction"],
                interval_15_str,
                interval_hour_str
            )

            self.compiled_data[key] += 1

    def write_compiled_data(self, export_path_excel):
        """
        Writes the aggregated data to an Excel file.

        Args:
            export_path_excel (str): Path to save the Excel report.
        
        Returns:
            str: Path to the saved Excel file.
        """
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Street Count Report"

        # Define headers
        headers = ['Site/Location', 'Date', 'Vehicle Type', 'Direction', 
                   '15 Min Interval', 'Hour Interval', 'Total Count']
        sheet.append(headers)

        # Sort keys for organized reporting
        sorted_keys = sorted(self.compiled_data.keys(), key=lambda x: (
            x[0], x[1], x[2], x[3], x[4], x[5]
        ))

        # Write data rows
        with logging_redirect_tqdm():
            for key in tqdm(sorted_keys, desc="Writing Rows", unit="rows"):
                row = list(key) + [self.compiled_data[key]]
                sheet.append(row)

        # Save the workbook
        workbook.save(export_path_excel)
        logging.info(f"Excel report saved at {export_path_excel}.")

        return export_path_excel
    
    def compile(self, output_path):
        """
        Executes the full compilation process.

        Args:
            export_path_excel (str): Path to save the Excel report.
        
        Returns:
            str: Path to the saved Excel file.
        """
        self.extract_data()
        self.precompile()
        return self.write_compiled_data(output_path)
